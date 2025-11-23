import asyncio
import csv
import json
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union, Iterator
from dataclasses import dataclass, field
from enum import Enum
import pandas as pd
import numpy as np
import aiohttp
import redis
import psycopg2
from psycopg2 import sql
from psycopg2.extras import RealDictCursor
import logging
from concurrent.futures import ThreadPoolExecutor
import threading
import queue
import time
import os
import tempfile
import zipfile
import gzip
import shutil
from pathlib import Path
import hashlib
import uuid
import base64
from contextlib import asynccontextmanager
from functools import wraps
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('export_service.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ExportFormat(Enum):
    JSON = "json"
    CSV = "csv"
    XML = "xml"
    EXCEL = "excel"
    PARQUET = "parquet"
    AVRO = "avro"
    ORC = "orc"
    HTML = "html"
    PDF = "pdf"
    MARKDOWN = "markdown"
    YAML = "yaml"

class ExportType(Enum):
    USERS = "users"
    CONVERSATIONS = "conversations"
    MESSAGES = "messages"
    ANALYTICS = "analytics"
    AUDIT_LOGS = "audit_logs"
    PERSONALIZATION_DATA = "personalization_data"
    SYSTEM_METRICS = "system_metrics"
    CUSTOM = "custom"

class ExportStatus(Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"

@dataclass
class ExportRequest:
    id: str
    user_id: str
    export_type: ExportType
    format: ExportFormat
    filters: Dict[str, Any] = field(default_factory=dict)
    fields: List[str] = field(default_factory=list)
    include_metadata: bool = True
    compression: bool = True
    chunk_size: int = 10000
    priority: int = 1
    scheduled_at: Optional[datetime] = None
    expires_at: Optional[datetime] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    created_at: datetime = field(default_factory=datetime.utcnow)
    updated_at: datetime = field(default_factory=datetime.utcnow)
    status: ExportStatus = ExportStatus.PENDING
    progress: float = 0.0
    file_path: Optional[str] = None
    file_size: Optional[int] = None
    download_url: Optional[str] = None
    error_message: Optional[str] = None

@dataclass
class ExportResult:
    request_id: str
    file_path: str
    file_size: int
    record_count: int
    export_duration: float
    compression_ratio: float
    checksum: str
    download_url: str
    metadata: Dict[str, Any] = field(default_factory=dict)

class DataExportService:
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or self._get_default_config()
        self.redis_client = None
        self.db_connection = None
        self.export_queue = queue.PriorityQueue()
        self.active_exports: Dict[str, ExportRequest] = {}
        self.export_history: List[ExportResult] = []
        self.executor = ThreadPoolExecutor(max_workers=5)
        self.is_running = False
        self.export_thread = None
        self.lock = threading.Lock()
        
        # Initialize connections
        self._initialize_connections()
        
        # Start export processor
        self.start_export_processor()
        
        logger.info("Data Export Service initialized successfully")

    def _get_default_config(self) -> Dict[str, Any]:
        """Get default configuration"""
        return {
            "database": {
                "host": "localhost",
                "port": 5432,
                "database": "chatbot",
                "user": "postgres",
                "password": "password"
            },
            "redis": {
                "host": "localhost",
                "port": 6379,
                "db": 0
            },
            "storage": {
                "base_path": "/tmp/exports",
                "max_file_size": "1GB",
                "retention_days": 30,
                "allowed_formats": ["json", "csv", "xml", "excel", "parquet"],
                "compression_formats": ["gzip", "zip"]
            },
            "export": {
                "max_concurrent_exports": 5,
                "default_chunk_size": 10000,
                "timeout": 3600,
                "retry_attempts": 3
            },
            "notifications": {
                "enabled": True,
                "email": {
                    "smtp_server": "localhost",
                    "smtp_port": 587,
                    "username": "exports@chatbot.com",
                    "password": "password"
                }
            }
        }

    def _initialize_connections(self):
        """Initialize database and Redis connections"""
        try:
            # Initialize Redis connection
            redis_config = self.config.get("redis", {})
            self.redis_client = redis.Redis(
                host=redis_config.get("host", "localhost"),
                port=redis_config.get("port", 6379),
                db=redis_config.get("db", 0),
                decode_responses=True
            )
            
            # Initialize database connection
            db_config = self.config.get("database", {})
            self.db_connection = psycopg2.connect(
                host=db_config.get("host", "localhost"),
                port=db_config.get("port", 5432),
                database=db_config.get("database", "chatbot"),
                user=db_config.get("user", "postgres"),
                password=db_config.get("password", "password"),
                cursor_factory=RealDictCursor
            )
            
            logger.info("Database and Redis connections established")
            
        except Exception as e:
            logger.error(f"Error initializing connections: {e}")
            raise

    def start_export_processor(self):
        """Start the export processor thread"""
        if not self.is_running:
            self.is_running = True
            self.export_thread = threading.Thread(target=self._export_processor)
            self.export_thread.daemon = True
            self.export_thread.start()
            logger.info("Export processor started")

    def stop_export_processor(self):
        """Stop the export processor thread"""
        self.is_running = False
        if self.export_thread:
            self.export_thread.join()
        logger.info("Export processor stopped")

    def _export_processor(self):
        """Process export requests from the queue"""
        while self.is_running:
            try:
                # Get next export request from queue
                priority, export_request = self.export_queue.get(timeout=1)
                
                with self.lock:
                    self.active_exports[export_request.id] = export_request
                
                # Process the export
                self._process_export(export_request)
                
                # Mark as completed
                with self.lock:
                    if export_request.id in self.active_exports:
                        del self.active_exports[export_request.id]
                
                self.export_queue.task_done()
                
            except queue.Empty:
                continue
            except Exception as e:
                logger.error(f"Error in export processor: {e}")
                # Mark export as failed
                if 'export_request' in locals():
                    export_request.status = ExportStatus.FAILED
                    export_request.error_message = str(e)
                    self._update_export_request(export_request)

    def _process_export(self, export_request: ExportRequest):
        """Process a single export request"""
        try:
            logger.info(f"Starting export {export_request.id} for user {export_request.user_id}")
            
            # Update status to processing
            export_request.status = ExportStatus.PROCESSING
            export_request.updated_at = datetime.utcnow()
            self._update_export_request(export_request)
            
            # Create temporary directory for export
            temp_dir = tempfile.mkdtemp()
            export_request.file_path = temp_dir
            
            # Get data based on export type
            data = self._get_export_data(export_request)
            
            # Export data to file
            file_path = self._export_to_file(data, export_request, temp_dir)
            
            # Calculate file size and checksum
            file_size = os.path.getsize(file_path)
            checksum = self._calculate_checksum(file_path)
            
            # Create download URL
            download_url = self._generate_download_url(export_request.id, file_path)
            
            # Create export result
            result = ExportResult(
                request_id=export_request.id,
                file_path=file_path,
                file_size=file_size,
                record_count=len(data) if isinstance(data, list) else 0,
                export_duration=(datetime.utcnow() - export_request.created_at).total_seconds(),
                compression_ratio=self._calculate_compression_ratio(file_path, data),
                checksum=checksum,
                download_url=download_url,
                metadata=export_request.metadata
            )
            
            # Update export request
            export_request.status = ExportStatus.COMPLETED
            export_request.file_path = file_path
            export_request.file_size = file_size
            export_request.download_url = download_url
            export_request.progress = 100.0
            export_request.updated_at = datetime.utcnow()
            
            # Store result
            with self.lock:
                self.export_history.append(result)
            
            # Update in database
            self._update_export_request(export_request)
            
            # Store in Redis for quick access
            self._store_export_result(result)
            
            logger.info(f"Export {export_request.id} completed successfully")
            
        except Exception as e:
            logger.error(f"Error processing export {export_request.id}: {e}")
            export_request.status = ExportStatus.FAILED
            export_request.error_message = str(e)
            self._update_export_request(export_request)

    def _get_export_data(self, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get data based on export type"""
        try:
            with self.db_connection.cursor() as cursor:
                if export_request.export_type == ExportType.USERS:
                    return self._get_users_data(cursor, export_request)
                elif export_request.export_type == ExportType.CONVERSATIONS:
                    return self._get_conversations_data(cursor, export_request)
                elif export_request.export_type == ExportType.MESSAGES:
                    return self._get_messages_data(cursor, export_request)
                elif export_request.export_type == ExportType.ANALYTICS:
                    return self._get_analytics_data(cursor, export_request)
                elif export_request.export_type == ExportType.AUDIT_LOGS:
                    return self._get_audit_logs_data(cursor, export_request)
                elif export_request.export_type == ExportType.PERSONALIZATION_DATA:
                    return self._get_personalization_data(cursor, export_request)
                elif export_request.export_type == ExportType.SYSTEM_METRICS:
                    return self._get_system_metrics_data(cursor, export_request)
                else:
                    return self._get_custom_data(cursor, export_request)
                    
        except Exception as e:
            logger.error(f"Error getting export data: {e}")
            raise

    def _get_users_data(self, cursor, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get users data for export"""
        query = "SELECT * FROM users"
        params = []
        
        # Apply filters
        if export_request.filters:
            conditions = []
            if "user_id" in export_request.filters:
                conditions.append("id = %s")
                params.append(export_request.filters["user_id"])
            if "email" in export_request.filters:
                conditions.append("email = %s")
                params.append(export_request.filters["email"])
            if "role" in export_request.filters:
                conditions.append("role = %s")
                params.append(export_request.filters["role"])
            if "created_after" in export_request.filters:
                conditions.append("created_at >= %s")
                params.append(export_request.filters["created_after"])
            if "created_before" in export_request.filters:
                conditions.append("created_at <= %s")
                params.append(export_request.filters["created_before"])
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
        
        # Apply field selection
        if export_request.fields:
            field_list = ", ".join(export_request.fields)
            query = query.replace("*", field_list)
        
        # Apply ordering
        query += " ORDER BY created_at DESC"
        
        # Apply pagination
        if "limit" in export_request.filters:
            query += " LIMIT %s"
            params.append(export_request.filters["limit"])
        if "offset" in export_request.filters:
            query += " OFFSET %s"
            params.append(export_request.filters["offset"])
        
        cursor.execute(query, params)
        return cursor.fetchall()

    def _get_conversations_data(self, cursor, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get conversations data for export"""
        query = """
            SELECT c.*, u.name as user_name, a.name as admin_name 
            FROM conversations c 
            LEFT JOIN users u ON c.user_id = u.id 
            LEFT JOIN users a ON c.admin_id = a.id
        """
        params = []
        
        # Apply filters
        if export_request.filters:
            conditions = []
            if "conversation_id" in export_request.filters:
                conditions.append("c.id = %s")
                params.append(export_request.filters["conversation_id"])
            if "user_id" in export_request.filters:
                conditions.append("c.user_id = %s")
                params.append(export_request.filters["user_id"])
            if "admin_id" in export_request.filters:
                conditions.append("c.admin_id = %s")
                params.append(export_request.filters["admin_id"])
            if "status" in export_request.filters:
                conditions.append("c.status = %s")
                params.append(export_request.filters["status"])
            if "started_after" in export_request.filters:
                conditions.append("c.started_at >= %s")
                params.append(export_request.filters["started_after"])
            if "started_before" in export_request.filters:
                conditions.append("c.started_at <= %s")
                params.append(export_request.filters["started_before"])
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
        
        # Apply field selection
        if export_request.fields:
            field_list = ", ".join(export_request.fields)
            query = query.replace("c.*, u.name as user_name, a.name as admin_name", field_list)
        
        # Apply ordering
        query += " ORDER BY c.started_at DESC"
        
        # Apply pagination
        if "limit" in export_request.filters:
            query += " LIMIT %s"
            params.append(export_request.filters["limit"])
        if "offset" in export_request.filters:
            query += " OFFSET %s"
            params.append(export_request.filters["offset"])
        
        cursor.execute(query, params)
        return cursor.fetchall()

    def _get_messages_data(self, cursor, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get messages data for export"""
        query = """
            SELECT m.*, c.title as conversation_title, u.name as sender_name 
            FROM messages m 
            LEFT JOIN conversations c ON m.conversation_id = c.id 
            LEFT JOIN users u ON m.sender_id = u.id
        """
        params = []
        
        # Apply filters
        if export_request.filters:
            conditions = []
            if "message_id" in export_request.filters:
                conditions.append("m.id = %s")
                params.append(export_request.filters["message_id"])
            if "conversation_id" in export_request.filters:
                conditions.append("m.conversation_id = %s")
                params.append(export_request.filters["conversation_id"])
            if "sender_id" in export_request.filters:
                conditions.append("m.sender_id = %s")
                params.append(export_request.filters["sender_id"])
            if "message_type" in export_request.filters:
                conditions.append("m.message_type = %s")
                params.append(export_request.filters["message_type"])
            if "created_after" in export_request.filters:
                conditions.append("m.created_at >= %s")
                params.append(export_request.filters["created_after"])
            if "created_before" in export_request.filters:
                conditions.append("m.created_at <= %s")
                params.append(export_request.filters["created_before"])
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
        
        # Apply field selection
        if export_request.fields:
            field_list = ", ".join(export_request.fields)
            query = query.replace("m.*, c.title as conversation_title, u.name as sender_name", field_list)
        
        # Apply ordering
        query += " ORDER BY m.created_at ASC"
        
        # Apply pagination
        if "limit" in export_request.filters:
            query += " LIMIT %s"
            params.append(export_request.filters["limit"])
        if "offset" in export_request.filters:
            query += " OFFSET %s"
            params.append(export_request.filters["offset"])
        
        cursor.execute(query, params)
        return cursor.fetchall()

    def _get_analytics_data(self, cursor, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get analytics data for export"""
        query = """
            SELECT 
                DATE_TRUNC('day', created_at) as date,
                COUNT(*) as total_users,
                COUNT(DISTINCT user_id) as active_users,
                COUNT(DISTINCT conversation_id) as total_conversations,
                COUNT(*) as total_messages,
                AVG(EXTRACT(EPOCH FROM (ended_at - started_at))) as avg_conversation_duration
            FROM conversations 
            WHERE created_at >= %s AND created_at <= %s
            GROUP BY DATE_TRUNC('day', created_at)
            ORDER BY date DESC
        """
        
        start_date = export_request.filters.get("start_date", datetime.utcnow() - timedelta(days=30))
        end_date = export_request.filters.get("end_date", datetime.utcnow())
        
        cursor.execute(query, (start_date, end_date))
        return cursor.fetchall()

    def _get_audit_logs_data(self, cursor, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get audit logs data for export"""
        query = "SELECT * FROM audit_logs"
        params = []
        
        # Apply filters
        if export_request.filters:
            conditions = []
            if "admin_id" in export_request.filters:
                conditions.append("admin_id = %s")
                params.append(export_request.filters["admin_id"])
            if "action" in export_request.filters:
                conditions.append("action = %s")
                params.append(export_request.filters["action"])
            if "created_after" in export_request.filters:
                conditions.append("created_at >= %s")
                params.append(export_request.filters["created_after"])
            if "created_before" in export_request.filters:
                conditions.append("created_at <= %s")
                params.append(export_request.filters["created_before"])
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
        
        # Apply ordering
        query += " ORDER BY created_at DESC"
        
        # Apply pagination
        if "limit" in export_request.filters:
            query += " LIMIT %s"
            params.append(export_request.filters["limit"])
        if "offset" in export_request.filters:
            query += " OFFSET %s"
            params.append(export_request.filters["offset"])
        
        cursor.execute(query, params)
        return cursor.fetchall()

    def _get_personalization_data(self, cursor, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get personalization data for export"""
        query = """
            SELECT up.*, u.name as user_name, u.email as user_email
            FROM user_profiles up
            LEFT JOIN users u ON up.user_id = u.id
        """
        params = []
        
        # Apply filters
        if export_request.filters:
            conditions = []
            if "user_id" in export_request.filters:
                conditions.append("up.user_id = %s")
                params.append(export_request.filters["user_id"])
            if "last_updated_after" in export_request.filters:
                conditions.append("up.last_updated >= %s")
                params.append(export_request.filters["last_updated_after"])
            if "last_updated_before" in export_request.filters:
                conditions.append("up.last_updated <= %s")
                params.append(export_request.filters["last_updated_before"])
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
        
        # Apply ordering
        query += " ORDER BY up.last_updated DESC"
        
        # Apply pagination
        if "limit" in export_request.filters:
            query += " LIMIT %s"
            params.append(export_request.filters["limit"])
        if "offset" in export_request.filters:
            query += " OFFSET %s"
            params.append(export_request.filters["offset"])
        
        cursor.execute(query, params)
        return cursor.fetchall()

    def _get_system_metrics_data(self, cursor, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get system metrics data for export"""
        query = """
            SELECT 
                metric_name,
                metric_value,
                timestamp,
                service_name,
                instance_id
            FROM system_metrics
            WHERE timestamp >= %s AND timestamp <= %s
            ORDER BY timestamp DESC
        """
        
        start_time = export_request.filters.get("start_time", datetime.utcnow() - timedelta(hours=24))
        end_time = export_request.filters.get("end_time", datetime.utcnow())
        
        cursor.execute(query, (start_time, end_time))
        return cursor.fetchall()

    def _get_custom_data(self, cursor, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Get custom data for export"""
        # Use custom query from metadata
        custom_query = export_request.metadata.get("custom_query")
        if not custom_query:
            raise ValueError("Custom query is required for custom export type")
        
        # Apply parameters
        params = export_request.metadata.get("query_params", [])
        
        cursor.execute(custom_query, params)
        return cursor.fetchall()

    def _export_to_file(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to file based on format"""
        try:
            if export_request.format == ExportFormat.JSON:
                return self._export_to_json(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.CSV:
                return self._export_to_csv(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.XML:
                return self._export_to_xml(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.EXCEL:
                return self._export_to_excel(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.PARQUET:
                return self._export_to_parquet(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.AVRO:
                return self._export_to_avro(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.ORC:
                return self._export_to_orc(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.HTML:
                return self._export_to_html(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.PDF:
                return self._export_to_pdf(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.MARKDOWN:
                return self._export_to_markdown(data, export_request, temp_dir)
            elif export_request.format == ExportFormat.YAML:
                return self._export_to_yaml(data, export_request, temp_dir)
            else:
                raise ValueError(f"Unsupported export format: {export_request.format}")
                
        except Exception as e:
            logger.error(f"Error exporting to file: {e}")
            raise

    def _export_to_json(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to JSON format"""
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.json")
        
        # Add metadata if requested
        if export_request.include_metadata:
            export_data = {
                "metadata": {
                    "export_id": export_request.id,
                    "export_type": export_request.export_type.value,
                    "format": export_request.format.value,
                    "exported_at": datetime.utcnow().isoformat(),
                    "record_count": len(data),
                    "filters": export_request.filters,
                    "fields": export_request.fields
                },
                "data": data
            }
        else:
            export_data = data
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_csv(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to CSV format"""
        if not data:
            file_path = os.path.join(temp_dir, f"export_{export_request.id}.csv")
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([])  # Empty CSV
            return file_path
        
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.csv")
        
        # Get field names
        fieldnames = export_request.fields if export_request.fields else list(data[0].keys())
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for row in data:
                # Filter fields if specified
                if export_request.fields:
                    filtered_row = {k: row.get(k) for k in export_request.fields}
                    writer.writerow(filtered_row)
                else:
                    writer.writerow(row)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_xml(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to XML format"""
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.xml")
        
        # Create root element
        root = ET.Element("export")
        
        # Add metadata
        if export_request.include_metadata:
            metadata = ET.SubElement(root, "metadata")
            ET.SubElement(metadata, "export_id").text = export_request.id
            ET.SubElement(metadata, "export_type").text = export_request.export_type.value
            ET.SubElement(metadata, "format").text = export_request.format.value
            ET.SubElement(metadata, "exported_at").text = datetime.utcnow().isoformat()
            ET.SubElement(metadata, "record_count").text = str(len(data))
            
            # Add filters
            filters = ET.SubElement(metadata, "filters")
            for key, value in export_request.filters.items():
                filter_elem = ET.SubElement(filters, "filter")
                filter_elem.set("key", key)
                filter_elem.text = str(value)
        
        # Add data
        data_elem = ET.SubElement(root, "data")
        
        for i, record in enumerate(data):
            record_elem = ET.SubElement(data_elem, "record")
            record_elem.set("id", str(i))
            
            for key, value in record.items():
                field_elem = ET.SubElement(record_elem, key)
                field_elem.text = str(value) if value is not None else ""
        
        # Write to file
        tree = ET.ElementTree(root)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_excel(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")
        
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.xlsx")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export Data"
        
        # Add metadata sheet if requested
        if export_request.include_metadata:
            metadata_ws = wb.create_sheet("Metadata")
            
            # Add metadata
            metadata_ws.append(["Export ID", export_request.id])
            metadata_ws.append(["Export Type", export_request.export_type.value])
            metadata_ws.append(["Format", export_request.format.value])
            metadata_ws.append(["Exported At", datetime.utcnow().isoformat()])
            metadata_ws.append(["Record Count", str(len(data))])
            metadata_ws.append([])
            
            # Add filters
            metadata_ws.append(["Filters:"])
            for key, value in export_request.filters.items():
                metadata_ws.append([f"{key}:", str(value)])
            
            # Add fields
            metadata_ws.append([])
            metadata_ws.append(["Fields:"])
            for field in export_request.fields:
                metadata_ws.append([field])
        
        # Add headers
        if data:
            headers = export_request.fields if export_request.fields else list(data[0].keys())
            ws.append(headers)
            
            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for record in data:
            if export_request.fields:
                row = [record.get(field) for field in export_request.fields]
            else:
                row = list(record.values())
            ws.append(row)
        
        # Auto-adjust column widths
        for col_num, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(file_path)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.zip'
            with zipfile.ZipFile(compressed_path, 'w') as zipf:
                zipf.write(file_path, os.path.basename(file_path))
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_parquet(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to Parquet format"""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            raise ImportError("pyarrow is required for Parquet export")
        
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.parquet")
        
        # Convert data to pandas DataFrame
        df = pd.DataFrame(data)
        
        # Select columns if specified
        if export_request.fields:
            df = df[export_request.fields]
        
        # Convert to Arrow Table
        table = pa.Table.from_pandas(df)
        
        # Write to Parquet file
        pq.write_table(table, file_path)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_avro(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to Avro format"""
        try:
            import fastavro
        except ImportError:
            raise ImportError("fastavro is required for Avro export")
        
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.avro")
        
        # Define Avro schema
        schema = {
            "type": "record",
            "name": "ExportRecord",
            "fields": [
                {"name": key, "type": ["null", "string"]} if isinstance(value, str) else
                {"name": key, "type": ["null", "int"]} if isinstance(value, int) else
                {"name": key, "type": ["null", "double"]} if isinstance(value, float) else
                {"name": key, "type": ["null", "boolean"]} if isinstance(value, bool) else
                {"name": key, "type": ["null", {"type": "array", "items": "string"}]} if isinstance(value, list) else
                {"name": key, "type": ["null", "string"]}  # Default to string
                for key, value in (data[0].items() if data else {}).items()
            ]
        }
        
        # Convert data to Avro format
        records = []
        for record in data:
            avro_record = {}
            for key, value in record.items():
                if value is not None:
                    avro_record[key] = value
                else:
                    avro_record[key] = None
            records.append(avro_record)
        
        # Write to Avro file
        with open(file_path, 'wb') as f:
            fastavro.writer(f, schema, records)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_orc(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to ORC format"""
        try:
            from pyarrow import orc
        except ImportError:
            raise ImportError("pyarrow is required for ORC export")
        
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.orc")
        
        # Convert data to pandas DataFrame
        df = pd.DataFrame(data)
        
        # Select columns if specified
        if export_request.fields:
            df = df[export_request.fields]
        
        # Convert to Arrow Table
        table = pa.Table.from_pandas(df)
        
        # Write to ORC file
        orc.write_table(table, file_path)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_html(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to HTML format"""
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.html")
        
        # Create HTML content
        html_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Export Data</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                table { border-collapse: collapse; width: 100%; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; font-weight: bold; }
                tr:nth-child(even) { background-color: #f9f9f9; }
                .metadata { background-color: #e7f3ff; padding: 15px; margin-bottom: 20px; border-radius: 5px; }
                .metadata h2 { margin-top: 0; }
                .metadata table { margin-bottom: 0; }
            </style>
        </head>
        <body>
        """
        
        # Add metadata
        if export_request.include_metadata:
            html_content += f"""
            <div class="metadata">
                <h2>Export Metadata</h2>
                <table>
                    <tr><td><strong>Export ID:</strong></td><td>{export_request.id}</td></tr>
                    <tr><td><strong>Export Type:</strong></td><td>{export_request.export_type.value}</td></tr>
                    <tr><td><strong>Format:</strong></td><td>{export_request.format.value}</td></tr>
                    <tr><td><strong>Exported At:</strong></td><td>{datetime.utcnow().isoformat()}</td></tr>
                    <tr><td><strong>Record Count:</strong></td><td>{len(data)}</td></tr>
                </table>
            </div>
            """
        
        # Add data table
        if data:
            html_content += "<h2>Export Data</h2>"
            html_content += "<table>"
            
            # Add headers
            headers = export_request.fields if export_request.fields else list(data[0].keys())
            html_content += "<thead><tr>"
            for header in headers:
                html_content += f"<th>{header}</th>"
            html_content += "</tr></thead>"
            
            # Add data rows
            html_content += "<tbody>"
            for record in data:
                html_content += "<tr>"
                if export_request.fields:
                    for field in export_request.fields:
                        value = record.get(field, "")
                        html_content += f"<td>{value}</td>"
                else:
                    for value in record.values():
                        html_content += f"<td>{value}</td>"
                html_content += "</tr>"
            html_content += "</tbody>"
            
            html_content += "</table>"
        
        html_content += "</body></html>"
        
        # Write to file
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_pdf(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to PDF format"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
        except ImportError:
            raise ImportError("reportlab is required for PDF export")
        
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.pdf")
        
        # Create PDF document
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        elements.append(Paragraph("Export Data", title_style))
        elements.append(Spacer(1, 12))
        
        # Add metadata
        if export_request.include_metadata:
            metadata_data = [
                ["Export ID", export_request.id],
                ["Export Type", export_request.export_type.value],
                ["Format", export_request.format.value],
                ["Exported At", datetime.utcnow().isoformat()],
                ["Record Count", str(len(data))]
            ]
            
            metadata_table = Table(metadata_data, colWidths=[2*inch, 3*inch])
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(metadata_table)
            elements.append(Spacer(1, 20))
        
        # Add data table
        if data:
            # Prepare data for table
            table_data = []
            
            # Add headers
            headers = export_request.fields if export_request.fields else list(data[0].keys())
            table_data.append(headers)
            
            # Add data rows
            for record in data:
                if export_request.fields:
                    row = [str(record.get(field, "")) for field in export_request.fields]
                else:
                    row = [str(value) for value in record.values()]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, colWidths=[1.5*inch] * len(headers))
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8)
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_markdown(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to Markdown format"""
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.md")
        
        markdown_content = []
        
        # Add metadata
        if export_request.include_metadata:
            markdown_content.append("# Export Metadata")
            markdown_content.append("")
            markdown_content.append(f"- **Export ID:** {export_request.id}")
            markdown_content.append(f"- **Export Type:** {export_request.export_type.value}")
            markdown_content.append(f"- **Format:** {export_request.format.value}")
            markdown_content.append(f"- **Exported At:** {datetime.utcnow().isoformat()}")
            markdown_content.append(f"- **Record Count:** {len(data)}")
            markdown_content.append("")
            
            if export_request.filters:
                markdown_content.append("## Filters")
                markdown_content.append("")
                for key, value in export_request.filters.items():
                    markdown_content.append(f"- **{key}:** {value}")
                markdown_content.append("")
        
        # Add data table
        if data:
            markdown_content.append("# Export Data")
            markdown_content.append("")
            
            # Prepare table headers
            headers = export_request.fields if export_request.fields else list(data[0].keys())
            
            # Add table headers
            markdown_content.append("| " + " | ".join(headers) + " |")
            markdown_content.append("|" + "|".join(["---"] * len(headers)) + "|")
            
            # Add table rows
            for record in data:
                if export_request.fields:
                    row = [str(record.get(field, "")) for field in export_request.fields]
                else:
                    row = [str(value) for value in record.values()]
                markdown_content.append("| " + " | ".join(row) + " |")
            
            markdown_content.append("")
        
        # Write to file
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(markdown_content))
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _export_to_yaml(self, data: List[Dict[str, Any]], export_request: ExportRequest, temp_dir: str) -> str:
        """Export data to YAML format"""
        try:
            import yaml
        except ImportError:
            raise ImportError("PyYAML is required for YAML export")
        
        file_path = os.path.join(temp_dir, f"export_{export_request.id}.yaml")
        
        # Add metadata if requested
        if export_request.include_metadata:
            export_data = {
                "metadata": {
                    "export_id": export_request.id,
                    "export_type": export_request.export_type.value,
                    "format": export_request.format.value,
                    "exported_at": datetime.utcnow().isoformat(),
                    "record_count": len(data),
                    "filters": export_request.filters,
                    "fields": export_request.fields
                },
                "data": data
            }
        else:
            export_data = data
        
        with open(file_path, 'w', encoding='utf-8') as f:
            yaml.dump(export_data, f, default_flow_style=False, allow_unicode=True)
        
        # Compress if requested
        if export_request.compression:
            compressed_path = file_path + '.gz'
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(file_path)
            return compressed_path
        
        return file_path

    def _calculate_checksum(self, file_path: str) -> str:
        """Calculate SHA256 checksum of a file"""
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)
        return sha256_hash.hexdigest()

    def _calculate_compression_ratio(self, file_path: str, original_data: List[Dict[str, Any]]) -> float:
        """Calculate compression ratio"""
        try:
            file_size = os.path.getsize(file_path)
            original_size = len(str(original_data).encode('utf-8'))
            return original_size / file_size if file_size > 0 else 1.0
        except:
            return 1.0

    def _generate_download_url(self, export_id: str, file_path: str) -> str:
        """Generate download URL for exported file"""
        # In a real implementation, this would generate a secure URL
        # For now, return a placeholder
        return f"/api/data/exports/{export_id}/download"

    def _store_export_result(self, result: ExportResult):
        """Store export result in Redis"""
        try:
            key = f"export_result:{result.request_id}"
            value = {
                "file_path": result.file_path,
                "file_size": result.file_size,
                "record_count": result.record_count,
                "export_duration": result.export_duration,
                "compression_ratio": result.compression_ratio,
                "checksum": result.checksum,
                "download_url": result.download_url,
                "metadata": result.metadata,
                "created_at": datetime.utcnow().isoformat()
            }
            
            self.redis_client.setex(key, 86400, json.dumps(value))  # 24 hours TTL
            
        except Exception as e:
            logger.error(f"Error storing export result: {e}")

    def _update_export_request(self, export_request: ExportRequest):
        """Update export request in database"""
        try:
            with self.db_connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE export_requests 
                    SET status = %s, progress = %s, file_path = %s, 
                        file_size = %s, download_url = %s, error_message = %s,
                        updated_at = %s
                    WHERE id = %s
                    """,
                    (export_request.status.value, export_request.progress,
                     export_request.file_path, export_request.file_size,
                     export_request.download_url, export_request.error_message,
                     export_request.updated_at, export_request.id)
                )
                self.db_connection.commit()
                
        except Exception as e:
            logger.error(f"Error updating export request: {e}")
            self.db_connection.rollback()

    def create_export_request(self, user_id: str, export_type: ExportType, format: ExportFormat,
                            filters: Dict[str, Any] = None, fields: List[str] = None,
                            include_metadata: bool = True, compression: bool = True,
                            chunk_size: int = 10000, priority: int = 1,
                            scheduled_at: datetime = None, expires_at: datetime = None,
                            metadata: Dict[str, Any] = None) -> str:
        """Create a new export request"""
        try:
            export_request = ExportRequest(
                id=str(uuid.uuid4()),
                user_id=user_id,
                export_type=export_type,
                format=format,
                filters=filters or {},
                fields=fields or [],
                include_metadata=include_metadata,
                compression=compression,
                chunk_size=chunk_size,
                priority=priority,
                scheduled_at=scheduled_at,
                expires_at=expires_at,
                metadata=metadata or {}
            )
            
            # Save to database
            with self.db_connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO export_requests 
                    (id, user_id, export_type, format, filters, fields, include_metadata, 
                     compression, chunk_size, priority, scheduled_at, expires_at, metadata, 
                     created_at, updated_at, status, progress)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (export_request.id, export_request.user_id, export_request.export_type.value,
                     export_request.format.value, json.dumps(export_request.filters),
                     json.dumps(export_request.fields), export_request.include_metadata,
                     export_request.compression, export_request.chunk_size, export_request.priority,
                     export_request.scheduled_at, export_request.expires_at,
                     json.dumps(export_request.metadata), export_request.created_at,
                     export_request.updated_at, export_request.status.value, export_request.progress)
                )
                self.db_connection.commit()
            
            # Add to queue if not scheduled for later
            if not scheduled_at or scheduled_at <= datetime.utcnow():
                self.export_queue.put((priority, export_request))
            
            logger.info(f"Created export request {export_request.id} for user {user_id}")
            return export_request.id
            
        except Exception as e:
            logger.error(f"Error creating export request: {e}")
            self.db_connection.rollback()
            raise

    def get_export_request(self, export_id: str) -> Optional[ExportRequest]:
        """Get export request by ID"""
        try:
            with self.db_connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT id, user_id, export_type, format, filters, fields, include_metadata,
                           compression, chunk_size, priority, scheduled_at, expires_at, metadata,
                           created_at, updated_at, status, progress, file_path, file_size, 
                           download_url, error_message
                    FROM export_requests WHERE id = %s
                    """,
                    (export_id,)
                )
                
                row = cursor.fetchone()
                if row:
                    return ExportRequest(
                        id=row['id'],
                        user_id=row['user_id'],
                        export_type=ExportType(row['export_type']),
                        format=ExportFormat(row['format']),
                        filters=json.loads(row['filters']) if row['filters'] else {},
                        fields=json.loads(row['fields']) if row['fields'] else [],
                        include_metadata=row['include_metadata'],
                        compression=row['compression'],
                        chunk_size=row['chunk_size'],
                        priority=row['priority'],
                        scheduled_at=row['scheduled_at'],
                        expires_at=row['expires_at'],
                        metadata=json.loads(row['metadata']) if row['metadata'] else {},
                        created_at=row['created_at'],
                        updated_at=row['updated_at'],
                        status=ExportStatus(row['status']),
                        progress=row['progress'],
                        file_path=row['file_path'],
                        file_size=row['file_size'],
                        download_url=row['download_url'],
                        error_message=row['error_message']
                    )
                return None
                
        except Exception as e:
            logger.error(f"Error getting export request: {e}")
            return None

    def get_export_status(self, export_id: str) -> Optional[Dict[str, Any]]:
        """Get export status by ID"""
        export_request = self.get_export_request(export_id)
        if not export_request:
            return None
        
        return {
            "id": export_request.id,
            "status": export_request.status.value,
            "progress": export_request.progress,
            "created_at": export_request.created_at.isoformat(),
            "updated_at": export_request.updated_at.isoformat(),
            "error_message": export_request.error_message
        }

    def cancel_export_request(self, export_id: str) -> bool:
        """Cancel an export request"""
        try:
            with self.db_connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE export_requests 
                    SET status = %s, updated_at = %s
                    WHERE id = %s AND status IN (%s, %s)
                    """,
                    (ExportStatus.CANCELLED.value, datetime.utcnow(), export_id,
                     ExportStatus.PENDING.value, ExportStatus.PROCESSING.value)
                )
                self.db_connection.commit()
                
                # Remove from queue if pending
                with self.lock:
                    items_to_remove = []
                    while not self.export_queue.empty():
                        priority, export_request = self.export_queue.get()
                        if export_request.id != export_id:
                            items_to_remove.append((priority, export_request))
                        else:
                            logger.info(f"Cancelled export request {export_id}")
                    
                    # Put back remaining items
                    for item in items_to_remove:
                        self.export_queue.put(item)
                
                return cursor.rowcount > 0
                
        except Exception as e:
            logger.error(f"Error cancelling export request: {e}")
            self.db_connection.rollback()
            return False

    def get_export_results(self, user_id: str = None, limit: int = 50, offset: int = 0) -> List[ExportResult]:
        """Get export results"""
        try:
            with self.db_connection.cursor() as cursor:
                if user_id:
                    cursor.execute(
                        """
                        SELECT request_id, file_path, file_size, record_count, 
                               export_duration, compression_ratio, checksum, download_url,
                               metadata, created_at
                        FROM export_results 
                        WHERE user_id = %s 
                        ORDER BY created_at DESC 
                        LIMIT %s OFFSET %s
                        """,
                        (user_id, limit, offset)
                    )
                else:
                    cursor.execute(
                        """
                        SELECT request_id, file_path, file_size, record_count, 
                               export_duration, compression_ratio, checksum, download_url,
                               metadata, created_at
                        FROM export_results 
                        ORDER BY created_at DESC 
                        LIMIT %s OFFSET %s
                        """,
                        (limit, offset)
                    )
                
                results = []
                for row in cursor.fetchall():
                    result = ExportResult(
                        request_id=row['request_id'],
                        file_path=row['file_path'],
                        file_size=row['file_size'],
                        record_count=row['record_count'],
                        export_duration=row['export_duration'],
                        compression_ratio=row['compression_ratio'],
                        checksum=row['checksum'],
                        download_url=row['download_url'],
                        metadata=json.loads(row['metadata']) if row['metadata'] else {}
                    )
                    results.append(result)
                
                return results
                
        except Exception as e:
            logger.error(f"Error getting export results: {e}")
            return []

    def get_export_formats(self) -> List[str]:
        """Get supported export formats"""
        return [format.value for format in ExportFormat]

    def get_export_types(self) -> List[str]:
        """Get supported export types"""
        return [export_type.value for export_type in ExportType]

    def cleanup_old_exports(self, days: int = 30):
        """Clean up old export files"""
        try:
            cutoff_date = datetime.utcnow() - timedelta(days=days)
            
            with self.db_connection.cursor() as cursor:
                # Get old export results
                cursor.execute(
                    """
                    SELECT request_id, file_path 
                    FROM export_results 
                    WHERE created_at < %s
                    """,
                    (cutoff_date,)
                )
                
                old_exports = cursor.fetchall()
                
                # Delete files
                for export in old_exports:
                    try:
                        if os.path.exists(export['file_path']):
                            if os.path.isdir(export['file_path']):
                                shutil.rmtree(export['file_path'])
                            else:
                                os.remove(export['file_path'])
                    except Exception as e:
                        logger.error(f"Error deleting file {export['file_path']}: {e}")
                
                # Delete from database
                cursor.execute(
                    """
                    DELETE FROM export_results 
                    WHERE created_at < %s
                    """,
                    (cutoff_date,)
                )
                
                self.db_connection.commit()
                logger.info(f"Cleaned up {len(old_exports)} old export files")
                
        except Exception as e:
            logger.error(f"Error cleaning up old exports: {e}")
            self.db_connection.rollback()

    def get_export_statistics(self) -> Dict[str, Any]:
        """Get export statistics"""
        try:
            with self.db_connection.cursor() as cursor:
                # Total exports
                cursor.execute("SELECT COUNT(*) FROM export_requests")
                total_exports = cursor.fetchone()['count']
                
                # Exports by status
                cursor.execute("""
                    SELECT status, COUNT(*) 
                    FROM export_requests 
                    GROUP BY status
                """)
                status_counts = {row['status']: row['count'] for row in cursor.fetchall()}
                
                # Exports by format
                cursor.execute("""
                    SELECT format, COUNT(*) 
                    FROM export_requests 
                    GROUP BY format
                """)
                format_counts = {row['format']: row['count'] for row in cursor.fetchall()}
                
                # Exports by type
                cursor.execute("""
                    SELECT export_type, COUNT(*) 
                    FROM export_requests 
                    GROUP BY export_type
                """)
                type_counts = {row['export_type']: row['count'] for row in cursor.fetchall()}
                
                # Average export duration
                cursor.execute("""
                    SELECT AVG(export_duration) 
                    FROM export_results 
                    WHERE export_duration IS NOT NULL
                """)
                avg_duration = cursor.fetchone()['avg']
                
                # Total data exported
                cursor.execute("""
                    SELECT SUM(record_count)
                    FROM export_results
                    WHERE record_count IS NOT NULL
                """)
                total_records = cursor.fetchone()['sum'] or 0
                
                return {
                    "total_exports": total_exports,
                    "status_counts": status_counts,
                    "format_counts": format_counts,
                    "type_counts": type_counts,
                    "average_export_duration": avg_duration,
                    "total_records_exported": total_records
                }
                
        except Exception as e:
            logger.error(f"Error getting export statistics: {e}")
            return {}

    def export_data_directly(self, export_type: ExportType, format: ExportFormat,
                           filters: Dict[str, Any] = None, fields: List[str] = None,
                           include_metadata: bool = True) -> str:
        """Export data directly and return file path"""
        try:
            # Create temporary directory
            temp_dir = tempfile.mkdtemp()
            
            # Get data
            with self.db_connection.cursor() as cursor:
                if export_type == ExportType.USERS:
                    data = self._get_users_data(cursor, ExportRequest(
                        id="", user_id="", export_type=export_type, format=format,
                        filters=filters or {}, fields=fields or [], include_metadata=include_metadata
                    ))
                elif export_type == ExportType.CONVERSATIONS:
                    data = self._get_conversations_data(cursor, ExportRequest(
                        id="", user_id="", export_type=export_type, format=format,
                        filters=filters or {}, fields=fields or [], include_metadata=include_metadata
                    ))
                elif export_type == ExportType.MESSAGES:
                    data = self._get_messages_data(cursor, ExportRequest(
                        id="", user_id="", export_type=export_type, format=format,
                        filters=filters or {}, fields=fields or [], include_metadata=include_metadata
                    ))
                else:
                    raise ValueError(f"Direct export not supported for type: {export_type}")
            
            # Export to file
            export_request = ExportRequest(
                id="", user_id="", export_type=export_type, format=format,
                filters=filters or {}, fields=fields or [], include_metadata=include_metadata
            )
            
            file_path = self._export_to_file(data, export_request, temp_dir)
            
            logger.info(f"Direct export completed: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error in direct export: {e}")
            raise

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.stop_export_processor()
        self.cleanup_old_exports()

# Example usage
if __name__ == "__main__":
    # Initialize export service
    with DataExportService() as service:
        
        # Create export request
        export_id = service.create_export_request(
            user_id="user123",
            export_type=ExportType.USERS,
            format=ExportFormat.JSON,
            filters={"created_after": datetime.utcnow() - timedelta(days=7)},
            fields=["id", "email", "name", "created_at"],
            include_metadata=True,
            compression=True
        )
        
        # Check export status
        status = service.get_export_status(export_id)
        print(f"Export status: {status}")
        
        # Get export results
        results = service.get_export_results(user_id="user123")
        print(f"Export results: {len(results)} files")
        
        # Get export statistics
        stats = service.get_export_statistics()
        print(f"Export statistics: {stats}")
        
        # Direct export
        file_path = service.export_data_directly(
            export_type=ExportType.CONVERSATIONS,
            format=ExportFormat.CSV,
            filters={"status": "active"},
            fields=["id", "user_id", "title", "started_at", "status"]
        )
        print(f"Direct export file: {file_path}")